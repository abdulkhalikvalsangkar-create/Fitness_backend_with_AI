"""Import the predefined FAQ set from an Excel or CSV sheet.

    python scripts/import_faq.py faqs.xlsx --dry-run     # inspect first
    python scripts/import_faq.py faqs.xlsx               # write
    python scripts/import_faq.py faqs.xlsx --sheet "Nutrition"
    python scripts/import_faq.py faqs.csv  --status draft

Column names are matched loosely, so a sheet written by a human does not have
to be reformatted before it can be loaded. "Question", "canonical question",
"FAQ", "Query" all resolve to the question column; "Answer", "Response",
"Reply" to the answer. Run with --dry-run first: it prints the mapping it
inferred and every row it would reject, and writes nothing.

Re-running is safe. Rows are upserted on a stable id derived from the question,
so correcting a typo in the sheet and re-importing updates that FAQ in place
rather than creating a second one. Surfaces are replaced wholesale per item,
so deleting a paraphrase from the sheet removes it from the database too.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator, Optional

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from packages.domain.enums import FaqCategory, FaqStatus, SafetyClass  # noqa: E402
from packages.domain.models import FaqItem, FaqVariants  # noqa: E402
from packages.storage.db import session_scope  # noqa: E402
from packages.storage.repositories.faq import FaqRepository  # noqa: E402

NOW = datetime.now(timezone.utc)

# Header synonyms, lowercased and stripped of non-letters before matching.
# Order matters only within a field: the first match wins.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "question": ("question", "canonicalquestion", "faq", "query", "prompt", "ask", "title"),
    "answer": ("answer", "answertemplate", "response", "reply", "solution", "content"),
    "category": ("category", "categories", "type", "topic", "section", "group"),
    "paraphrases": (
        "paraphrases", "paraphrase", "synonyms", "synonym", "variations", "variants",
        "alternatequestions", "alternates", "similarquestions", "keywords", "tags",
    ),
    "safety": ("safety", "safetyclass", "sensitivity", "risk"),
    "id": ("id", "faqid", "key", "code", "slug"),
    "locale": ("locale", "language", "lang"),
    "status": ("status", "state", "published"),
    "answer_with_data": ("answerwithdata", "withdata", "personalisedanswer", "personalizedanswer"),
    "answer_without_data": ("answerwithoutdata", "withoutdata", "genericanswer", "fallback"),
    "required_slots": ("requiredslots", "slots", "requiredfields", "needs"),
    "owner": ("owner", "author", "createdby", "team"),
}

CATEGORY_ALIASES: dict[str, FaqCategory] = {
    "general": FaqCategory.GENERAL,
    "nutrition": FaqCategory.NUTRITION,
    "diet": FaqCategory.NUTRITION,
    "food": FaqCategory.NUTRITION,
    "workout": FaqCategory.WORKOUT,
    "exercise": FaqCategory.WORKOUT,
    "training": FaqCategory.WORKOUT,
    "fitness": FaqCategory.WORKOUT,
    "medical": FaqCategory.MEDICAL,
    "health": FaqCategory.MEDICAL,
    "clinical": FaqCategory.MEDICAL,
    "product": FaqCategory.PRODUCT,
    "products": FaqCategory.PRODUCT,
    "appsupport": FaqCategory.APP_SUPPORT,
    "app": FaqCategory.APP_SUPPORT,
    "support": FaqCategory.APP_SUPPORT,
    "technical": FaqCategory.APP_SUPPORT,
    "account": FaqCategory.APP_SUPPORT,
}

SAFETY_ALIASES: dict[str, SafetyClass] = {
    "informational": SafetyClass.INFORMATIONAL,
    "info": SafetyClass.INFORMATIONAL,
    "general": SafetyClass.INFORMATIONAL,
    "guidance": SafetyClass.GUIDANCE,
    "advice": SafetyClass.GUIDANCE,
    "medical": SafetyClass.MEDICAL_SENSITIVE,
    "medicalsensitive": SafetyClass.MEDICAL_SENSITIVE,
    "sensitive": SafetyClass.MEDICAL_SENSITIVE,
    "high": SafetyClass.MEDICAL_SENSITIVE,
}

# One cell holding several paraphrases — newline, pipe, or semicolon. Commas
# are NOT a separator: real questions contain them ("how much protein, roughly?")
# and splitting on them shreds the paraphrase into fragments that match nothing.
_PARAPHRASE_SPLIT = re.compile(r"[\n|;]+")

_SLUG_STRIP = re.compile(r"[^a-z0-9]+")


def _norm_header(value: Any) -> str:
    return _SLUG_STRIP.sub("", str(value or "").strip().lower())


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").strip()
    return "" if text.lower() in ("nan", "none", "null", "-") else text


def make_faq_id(question: str, explicit: str = "") -> str:
    """Stable id so re-importing updates rather than duplicates.

    A readable slug plus a short hash: the slug makes rows recognisable in the
    database, the hash keeps two similarly-worded questions apart.
    """
    if explicit:
        slug = _SLUG_STRIP.sub("_", explicit.strip().lower()).strip("_")
        if slug:
            return f"faq_{slug}"[:64]

    words = _SLUG_STRIP.sub(" ", question.lower()).split()[:6]
    slug = "_".join(words) or "item"
    digest = hashlib.sha256(question.strip().lower().encode()).hexdigest()[:8]
    return f"faq_{slug}_{digest}"[:64]


def detect_columns(headers: list[str]) -> dict[str, int]:
    """Map our field names to column indexes. Unmatched fields are absent."""
    normalised = [_norm_header(h) for h in headers]
    mapping: dict[str, int] = {}

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in normalised:
                mapping[field] = normalised.index(alias)
                break
        else:
            # Fall back to a prefix match: "question (english)" -> question.
            for index, header in enumerate(normalised):
                if header and any(header.startswith(a) for a in aliases):
                    if index not in mapping.values():
                        mapping[field] = index
                        break
    return mapping


def read_rows(path: Path, sheet: Optional[str]) -> tuple[list[str], Iterator[list[Any]]]:
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover
            raise SystemExit(
                "reading .xlsx needs openpyxl:  pip install openpyxl\n"
                "(or export the sheet as CSV and pass that instead)"
            ) from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [_clean(h) for h in next(rows)]
        except StopIteration:
            raise SystemExit(f"{path.name} is empty")
        return headers, (list(r) for r in rows)

    if path.suffix.lower() == ".xls":
        raise SystemExit(
            "legacy .xls is not supported — open it in Excel and save as .xlsx or .csv"
        )

    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = list(csv.reader(handle))
    if not reader:
        raise SystemExit(f"{path.name} is empty")
    return [_clean(h) for h in reader[0]], iter(reader[1:])


def build_item(
    row: list[Any],
    columns: dict[str, int],
    default_status: FaqStatus,
    default_locale: str,
    owner: str,
) -> tuple[Optional[FaqItem], str]:
    """Returns (item, problem). Exactly one of the two is set."""

    def cell(field: str) -> str:
        index = columns.get(field)
        if index is None or index >= len(row):
            return ""
        return _clean(row[index])

    question = cell("question")
    answer = cell("answer")

    if not question and not answer:
        return None, ""  # blank row — skipped silently, not an error
    if not question:
        return None, "no question"
    if not answer:
        return None, f"no answer for {question[:60]!r}"

    raw_category = _norm_header(cell("category"))
    category = CATEGORY_ALIASES.get(raw_category, FaqCategory.GENERAL)

    raw_safety = _norm_header(cell("safety"))
    if raw_safety:
        safety = SAFETY_ALIASES.get(raw_safety, SafetyClass.INFORMATIONAL)
    else:
        # No safety column: a medical question is the one place where guessing
        # wrong matters, so it inherits the stricter class from its category.
        safety = (
            SafetyClass.MEDICAL_SENSITIVE
            if category is FaqCategory.MEDICAL
            else SafetyClass.INFORMATIONAL
        )

    paraphrases = [
        p.strip()
        for p in _PARAPHRASE_SPLIT.split(cell("paraphrases"))
        if p.strip() and p.strip().lower() != question.lower()
    ]

    slots = [
        s.strip()
        for s in re.split(r"[,\n|;]+", cell("required_slots"))
        if s.strip()
    ]

    with_data = cell("answer_with_data")
    without_data = cell("answer_without_data")
    variants = (
        FaqVariants(with_data=with_data or None, without_data=without_data or None)
        if (with_data or without_data)
        else FaqVariants()
    )

    raw_status = _norm_header(cell("status"))
    status = default_status
    if raw_status in ("live", "published", "active", "yes", "true", "1"):
        status = FaqStatus.LIVE
    elif raw_status in ("draft", "no", "false", "0", "pending"):
        status = FaqStatus.DRAFT

    return (
        FaqItem(
            id=make_faq_id(question, cell("id")),
            status=status,
            category=category,
            canonical_question=question,
            paraphrases=paraphrases,
            answer_template=answer,
            variants=variants,
            required_slots=slots,
            safety_class=safety,
            locale=cell("locale") or default_locale,
            owner=cell("owner") or owner,
            reviewed_by=cell("owner") or owner,
            reviewed_at=NOW,
        ),
        "",
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("path", help="path to the .xlsx or .csv sheet")
    parser.add_argument("--sheet", default=None, help="worksheet name (default: first)")
    parser.add_argument("--dry-run", action="store_true", help="report only; write nothing")
    parser.add_argument(
        "--status",
        choices=["live", "draft"],
        default="live",
        help="status for rows whose sheet does not say (default: live)",
    )
    parser.add_argument("--locale", default="en")
    parser.add_argument("--owner", default="import")
    args = parser.parse_args()

    path = Path(args.path)
    if not path.is_file():
        print(f"ERROR: no such file: {path}")
        return 2

    headers, rows = read_rows(path, args.sheet)
    columns = detect_columns(headers)

    print(f"file    : {path.name}")
    print(f"headers : {headers}")
    print("\ncolumn mapping")
    for field in COLUMN_ALIASES:
        index = columns.get(field)
        # ASCII only: Windows consoles default to cp1252 and turn a dash into
        # a literal '?', which reads like the tool is unsure about the mapping.
        shown = f"{headers[index]!r} (col {index + 1})" if index is not None else "(not found)"
        print(f"  {field:<22} {shown}")

    if "question" not in columns or "answer" not in columns:
        print(
            "\nERROR: could not find a question and an answer column.\n"
            "Rename them to 'Question' and 'Answer', or tell me the headers "
            "you use and I'll add them as aliases."
        )
        return 2

    default_status = FaqStatus.LIVE if args.status == "live" else FaqStatus.DRAFT

    items: list[FaqItem] = []
    problems: list[str] = []
    seen_ids: dict[str, str] = {}

    for number, row in enumerate(rows, start=2):
        item, problem = build_item(row, columns, default_status, args.locale, args.owner)
        if problem:
            problems.append(f"row {number}: {problem}")
            continue
        if item is None:
            continue
        if item.id in seen_ids:
            # Two rows with the same question: the second would silently
            # overwrite the first, so say so rather than losing content.
            problems.append(
                f"row {number}: duplicate of row {seen_ids[item.id]} "
                f"({item.canonical_question[:50]!r})"
            )
            continue
        seen_ids[item.id] = str(number)
        items.append(item)

    by_category: dict[str, int] = {}
    surfaces = 0
    for item in items:
        by_category[item.category.value] = by_category.get(item.category.value, 0) + 1
        surfaces += len(item.paraphrases) + 1

    print(f"\nparsed  : {len(items)} FAQ item(s), {surfaces} surface form(s)")
    for name in sorted(by_category):
        print(f"    {name:<12} {by_category[name]}")

    if problems:
        print(f"\nskipped : {len(problems)} row(s)")
        for line in problems[:20]:
            print(f"    {line}")
        if len(problems) > 20:
            print(f"    … and {len(problems) - 20} more")

    if items:
        sample = items[0]
        print("\nfirst item preview")
        print(f"    id        {sample.id}")
        print(f"    category  {sample.category.value}   safety {sample.safety_class.value}")
        print(f"    question  {sample.canonical_question[:90]}")
        print(f"    answer    {sample.answer_template[:90]}")
        print(f"    surfaces  {[sample.canonical_question[:40]] + sample.paraphrases[:3]}")

    if args.dry_run:
        print("\n--dry-run: nothing written.")
        return 0

    if not items:
        print("\nnothing to import.")
        return 1

    written = 0
    total_surfaces = 0
    with session_scope() as session:
        repo = FaqRepository(session)
        for item in items:
            repo.upsert_item(item)
            total_surfaces += repo.replace_surfaces(item.id, item)
            written += 1

    print(f"\nimported {written} FAQ item(s), {total_surfaces} surface form(s).")
    print("Exact (S1) and fulltext (S2) routing use these immediately; no backfill needed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
